# -*- coding: utf-8 -*-
import json, os, datetime, sqlite3, uuid, io
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment, colors
from openpyxl.cell import Cell
from openpyxl import Workbook

import boto3
from boto3.dynamodb.conditions import Key

s3 = boto3.resource('s3')
dynamodb = boto3.resource('dynamodb')
TABLE = os.environ.get("TABLE")
ENV = os.environ.get("ENV")
table = dynamodb.Table(TABLE)
Bucket = os.environ.get("BUCKET")


############################
# GENERATING EXCEL EXPORTS #
############################

def makeResponse(res, code):
    return {
        "statusCode": code,
        "body": json.dumps(res),
        "headers": {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Credentials': True,
            'Content-Type' : 'application/json'
        }
    }


def upload_post(event, context):
    body = json.loads(event["body"])
    upload_id = event["pathParameters"]["upload_id"]
    if upload_id == 'start':
        upload_id = uuid.uuid4()
        initUpload(upload_id, body)
        return makeResponse(str(upload_id), 200) 
    else:
        addTotables(upload_id, body)
        return makeResponse(upload_id, 200) 

def upload_get(event, context):
    upload_id = event["pathParameters"]["upload_id"]
    url = createWb(upload_id)
    print(url)
    return makeResponse({"url":url}, 200)

# Convert Hexadecimal color to RGB
def hexToRgb(val):
    h = val.lstrip('#')
    if len(h) == 3:
        h =h[0]+h[0]+h[1]+h[1]+h[2]+h[2]
    return "00" + h

# Create a Matrix
def addToMatrix(m, col_i, row_j, val):
    len_rows = len(m)
    len_cols = 0 if len_rows == 0 else len(m[0])
    if row_j > len_rows:
        for k in range(row_j - len_rows):
            m.append([None] * len_cols)
    if col_i > len_cols:
        for row in m:
            row += [None] * (col_i - len_cols)
    
    m[row_j - 1][col_i - 1] = val

# Extract Metadata from frontend dictionary
def getMeta(meta):
    with_headers_row = meta["with_headers_row"]
    with_headers_col = meta["with_headers_col"]
    with_alias_row = meta.get("with_alias_row", False)
    sorting = meta["sorting"]
    rows = meta["rows"]
    columns = meta["columns"]
    measures = meta["measures"]
    headers = [_["name"] for _ in meta["headers"]]
    merge = meta.get("merge", True)
    if len(sorting) == 0:
        sorting = {_:True for _ in rows+columns}
    else:
        sorting = {_["name"]:_["sorted"] if 'sorted' in _.keys() else meta["custom_sort"][_["name"]] for _ in sorting}
    return with_headers_row, with_headers_col, with_alias_row, sorting, rows, columns, measures, merge, headers

# Generate the tree data structure for the Rows/Columns
# (iterative)
def insertTree(tree, val, header):
    if val["_formattedValue"] not in [_["name"] for _ in tree["children"]]: tree["children"].append({"name":val["_formattedValue"], "value": val["_value"] if str(val["_value"])[0] != "[" else val["_formattedValue"], "header" :header, "children":[]})
    return tree["children"][[_["name"] for _ in tree["children"]].index(val["_formattedValue"])]

# Measures each branch of the tree to know how many rows/columns it contains
# (iterative)
def measureTree(tree):
    i = 0
    for child in tree["children"]:
        i+=measureTree(child)
    i = 1 if i == 0 else i
    tree["length"] = i
    return i

# Try sort
def trySort(arr, reverse):
    try:
        arr = sorted(arr, key=sortFunc, reverse=reverse)
    except:
        arr = sorted(arr, key=sortFuncStr, reverse=reverse)
    return arr

# Sort each branch of the tree alphabetically or reverse-alphabetically or manually
def sortTree(tree, sorting):
    if len(tree["children"]) > 0:
        if isinstance(sorting[tree["children"][0]["header"]], dict):
            compare_list = sorting[tree["children"][0]["header"]]["sort"]

            new_elements = []
            sort_indexes = []
            sort_elements = []

            for x in tree["children"]:
                if x["name"] in compare_list:
                    sort_indexes.append(compare_list.index(x["name"]))
                    sort_elements.append(x)
                else:
                    new_elements.append(x)

            sort_elements = [x for _,x in sorted(zip(sort_indexes, sort_elements))]

            new_elements = trySort(new_elements, sorting[tree["children"][0]["header"]]["sorted"])

            tree["children"] = sort_elements + new_elements
        else:
            tree["children"] = trySort(tree["children"], sorting[tree["children"][0]["header"]])
        
        for child in tree["children"]:
            sortTree(child, sorting)

# Small function used by the sortTree function to sort on the value attribute
sortFunc = lambda a : a["value"].upper() if type(a["value"]) is str else a["value"]
sortFuncStr = lambda a: str(a["value"]).upper()

# Mark the start and end of Rows/Columns in the data tree structure
def countTree(tree, start):
    for child in tree["children"]:
        child["start"] = start
        countTree(child, start)
        start = start + child["length"]
        child["end"] = start
        
# Extract colors and fonts from the raw dictionary
# - separating fonts from background colors
def getColors(colors):
    output = {}
    font = {}
    for key, value in colors.items():
        if "border" in key:
            border = hexToRgb(value["hex"])
            output[key] = Border(left=Side(style='thin', color=border), right=Side(style='thin', color=border), top=Side(style='thin', color=border), bottom=Side(style='thin', color=border))
        else:
            fill = hexToRgb(value["hex"])
            font[key] = invertColor(value["hex"])
            output[key] = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
    return output, font

# Accessing the child of a tree element
def getChild(tree, val):
    if len(tree["children"]) == 0:
        return tree
    else:
        i = [_["name"] for _ in tree["children"]].index(val["_formattedValue"])
        return tree["children"][i]

# Applying the extremes according to current rows and columns
# - this is used to fill in empty cells (with no data) from minimum row/col to maximum row/col
def setExtremes(extremes, r, c):
    extremes["rows"][0] = min(extremes["rows"][0], r+1)
    extremes["rows"][1] = max(extremes["rows"][1], r+1)
    extremes["cols"][0] = min(extremes["cols"][0], c+1)
    extremes["cols"][1] = max(extremes["cols"][1], c+1)

# Writes row aliases
def writeAlias(ws, aliases, start_row, start_col, colors=None, fonts=None, has_styling=True):
    for i in range(len(aliases)):
        cell = ws.cell(row=start_row+1, column=start_col+1+i)
        try:
            if has_styling:
                cell.fill = colors["header_fill"]
                cell.font = fonts["header_fill"]
                cell.border = colors["header_border"]
            cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.value = aliases[i]
        except:
            pass
        


# Writing Column/row headers according to colors selected
# - will merge columns like in Tableau if option was selected by the user
def writeHeaders(ws, tree, _type, start=0, colors=None, fonts=None, merge=True, has_styling=True):
    horizontal = "center" if _type != "rows" else "left"
    if "start" in tree.keys() and "end" in tree.keys():
        extremes = {"rows" : [100000, 0], "cols":[100000, 0]}
        for i in range(tree["start"], tree["end"]):
            r = i if _type  == "rows" else start - 1
            c = i if _type  != "rows" else start - 1
            setExtremes(extremes, r, c)
            cell = ws.cell(row=r+1, column=c+1)
            try:
                if has_styling:
                    cell.fill = colors["header_fill"]
                    cell.font = fonts["header_fill"]
                    cell.border = colors["header_border"]
                cell.value = tree["name"]
                cell.alignment = Alignment(horizontal=horizontal, vertical='top')
            except:
                pass
            
        if merge and moreThanOneCell(extremes, _type):
            ws.merge_cells(start_row=extremes["rows"][0], start_column=extremes["cols"][0], end_row=extremes["rows"][1], end_column=extremes["cols"][1])

    for child in tree["children"]:
        writeHeaders(ws, child, _type, start=start+1, colors=colors, fonts=fonts, merge=merge, has_styling=has_styling)

def moreThanOneCell(extremes, _type):
    if _type == "rows":
        return extremes["rows"][0] != extremes["rows"][1]
    else:
        return extremes["cols"][0] != extremes["cols"][1]

# Adds zeros in front of a number 
# - not used
def padZero(_str, l=2):
    zeros = ['0' for _ in range(l)]
    return (zeros + _str)[-l:]

# Returns a black or white color opposite to the hexadecimal color provided
# - will return black if the background color is slightly light and white if otherwise
# - used for font colors
def invertColor(_hex):
    if _hex[0] == "#":
        _hex = _hex[1:]
    if len(_hex) == 3:
        _hex = "".join([_+_ for _ in _hex])
    if len(_hex) != 6:
        raise 'Invalid HEX color.'
    r = int(_hex[:2], base=16)
    g = int(_hex[2:4], base=16)
    b = int(_hex[4:], base=16)
    return Font(color=colors.BLACK) if (r * 0.299 + g * 0.587 + b * 0.114) > 186 else Font(color=colors.WHITE) 

# Fills the empty colors thanks to extremes rows/columns
def finishFilling(ws, extremes, colors, fonts):
    for c in range(extremes["cols"][0], extremes["cols"][1]+1):
        for r in range(extremes["rows"][0], extremes["rows"][1]+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = colors["pane_fill_0"] if (r+1)%2 == 0 else colors["pane_fill_1"]
            cell.font   = fonts["pane_fill_0"]  if (r+1) % 2 == 0 else fonts["pane_fill_1"]
            cell.border = colors["pane_border"]

# Extract header
def extractHeader(t, output =[]):
    if "header" in t.keys():
        output.append(t["header"])
    if "children" in t.keys() and len(t["children"]) > 0:
        return extractHeader(t["children"][0], output)
    else:
        return output

# Return Connection
def returnConn():
    return sqlite3.connect('excelExtract.db')

def extractBody(body):
    desc = "Here is the Excel data extract from your Tableau Dashboard"
    title = body.get("title", "Dashboard")
    disclaimer_name = body.get("disclaimer_name", "")
    description =body.get("description", desc)
    mode = body.get("mode", "many")
    orientation = body.get("orientation", "vertical")
    has_styling = body.get("has_styling", "true")
    metadata = json.dumps(body["metadata"])
    return title, disclaimer_name, description, mode, orientation, has_styling, metadata

# Main initialization
def initUpload(_id, body):
    title, disclaimer_name, description, mode, orientation, has_styling, metadata = extractBody(body)
    table.put_item(Item={
        'uuid': str(_id),
        'table_id':-1,
        'title': title,
        'disclaimer_name': disclaimer_name,
        'description': description,
        'mode': mode,
        'orientation':orientation,
        'has_styling': has_styling,
        'metadata':metadata
    })

def flush(uuid, max_id):
    for _id in range(-1, max_id):
        table.delete_item(Key={
            'uuid': uuid,
            'table_id': _id
        })

# Adding to tables
def addTotables(uuid, data):
    table.put_item(Item={
        'uuid': uuid,
        'table_id': int(data["upload_id"]),
        'metadata': json.dumps(data)
    })

def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

def getMetadata(uuid):
    print("getting metadata", uuid)
    tables = table.query(
        KeyConditionExpression=Key('uuid').eq(uuid)
    )
    flush(uuid, len(tables["Items"]))
    res = {}
    output = []
    for t in tables["Items"]:
        if t["table_id"] == -1:
            res = t
        else:
            el = json.loads(t["metadata"])
            output.append(el)
    return res, output
      
# Main function called by the POST HTTP call
def createWb(uid, excel_path="./"):
    # Initialize the offset
    offset_col = 0
    offset_row = 0

    _h = {}

    r, tables = getMetadata(uid)
    print(r, len(tables))

    # Extracting all the information necessary
    metadata = json.loads(r["metadata"])
    title = r.get("title", "Dashboard")
    disclaimer_name = r.get("disclaimer_name", None)
    description =r.get("description", "Here is the Excel data extract from your Tableau Dashboard")
    has_styling = r.get("has_styling", "true") == "true"
    mode = r.get("mode", "many")
    orientation = r.get("orientation", "vertical") #used to check space

    # Creating an Excel Workbook from scratch
    wb = Workbook()
    ws = None

    # Creating the disclaimer page if there is one
    u = 0
    first_page = True
    if disclaimer_name != None:
        ws = wb.create_sheet(disclaimer_name, 0)
        cell = ws.cell(row=2, column=2)
        cell.value = description
        cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions["B"].width = 150
        ws.row_dimensions[2].height = 150
        u += 1
    # End of creation of the disclaimer page
    _matrix = []
    # There can be many tables exported at once, iterating through them
    
    for tab in tables:
        # Extracting the data retrieved by the extension
        data = tab["_data"]

        # headers = [_["_fieldName"] for _ in tab["_columns"]] 
        headers = []
        for c in tab["_columns"]:
            _ = c["_fieldName"]
            # _ =  _[4:-1] if _.startswith("AGG(") and _.endswith(")") else _
            headers.append(_)

        value = tab["alias"]
        key = tab["name"]

        margin_col = int(tab["margin_col"])
        margin_row = int(tab["margin_row"])

        # Initializing Extremes
        extremes = {"rows" : [100000, 0], "cols":  [100000, 0]}

        # If the mode is set to "many", then each tab should have its own sheet
        if mode == "many" or first_page: 
            ws = wb.create_sheet(value, u)
            first_page = False

        # Extracting colors
        colors, fonts = getColors(tab.get("color", {
            "header_border": { "hex": "#C40F25" }, "header_fill" : { "hex": "#C40F25" },
            "pane_border"  : { "hex": "#ffffff" }, "pane_fill_0" : { "hex": "#9B9B9B" },
            "pane_fill_1"  : { "hex": "#eeeeee" }, "null"        : { "hex": "#C40F25" }
        }))

        # Extracting metadata for the current table
        with_headers_row, with_headers_col, with_alias_row, sorting, rows, columns, measures, merge, headers_x = getMeta(metadata[key])

        # Initializing the data tree structure
        row_tree = {"name":"root", "value":"root", "children":[]}
        col_tree = {"name":"root", "value":"root", "children":[]}

        # Iterating throught the data to populate the data tree structure
        for row in data:
            row_o = row_tree
            col_o = col_tree
            for col in columns:
                i = col["index"]
                val = row[i]
                col_o = insertTree(col_o, val, headers[i])
            for r in rows:
                i = r["index"]
                val = row[i]
                row_o = insertTree(row_o, val, headers[i])
        # Measure the number of rows (start and end)
        measureTree(row_tree)
        # Measure the number of columns (start and end)
        measureTree(col_tree)

        # Sorting the rows according to the sorting selected
        sortTree(row_tree, sorting)
        # Sorting the columns according to the sorting selected
        sortTree(col_tree, sorting)
        lst = []

        # Applying the offset row/column
        # It basically adds an offset if the user selected one or if another table 
        # had already been inserted in the current sheet
        start_offset_row = (len(columns)+offset_row if with_headers_col else offset_row) + margin_row
        start_offset_col = (len(rows)+offset_col if with_headers_row else offset_col) + margin_col
        
        # Counts the rows and columns
        countTree(row_tree, start_offset_row)
        countTree(col_tree, start_offset_col)

        count = 0

        # Adds headers if selected by the user
        if with_headers_row:
            if with_alias_row:
                # write the alias for rows
                _h[key] = extractHeader(row_tree, output=[])
                writeAlias(ws, _h[key], start_offset_row-1, margin_col+offset_col, colors=colors, fonts=fonts, has_styling=has_styling)
            writeHeaders(ws, row_tree, "rows", start=margin_col+offset_col, colors=colors, fonts=fonts, merge=merge, has_styling=has_styling)
           
            if with_headers_col:
                writeHeaders(ws, col_tree, "columns", start=margin_row+offset_row, colors=colors, fonts=fonts, merge=merge, has_styling=has_styling)
        
        # fills in the actual data where it should be according to sorting and cells offsets
        if len(measures) != 0:
            for row in data:
                count+=1
                c = col_tree
                r = row_tree
                for col in columns:
                    i = col["index"]
                    c = getChild(c, row[i])
                for ro in rows:
                    i = ro["index"]
                    r = getChild(r, row[i])

                vals = []
                als = []
                for me in measures:
                    i = me["index"]
                    print(i, vals, row)
                    try:
                        vals.append(row[i]["_value"])
                        als.append(row[i]["_formattedValue"])
                    except:
                        pass
                if "start" not in r.keys(): r["start"] = start_offset_row
                if "start" not in c.keys(): c["start"] = start_offset_col
                cell = ws.cell(row=r["start"]+1, column=c["start"]+1)

                # Updating min/max for columns and rows for the current table
                setExtremes(extremes, r["start"], c["start"])
                
                try:
                    if als[0].endswith("%"):
                        cell.number_format = '0.00%'
                    cell.value  = vals[0]
                except:
                    pass

        # Offset has to be updated if we are going to change of tables
        # we set the offset according to the table orientation and the found extremes
        if mode != "many":
            if orientation == "vertical":
                offset_row = extremes["rows"][1]
            else:
                offset_col = extremes["cols"][1]

        # we finish to fill in the colors even for empty cells
        if has_styling:
            finishFilling(ws, extremes, colors, fonts)

        u+=1
    # we add a datetime tag in the filename
    d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M-%S")
    n = title+"_"+d+".xlsx"
    lst =wb.get_sheet_names()
    # we potentially remove the standard Sheet which is empty
    if "Sheet" in lst and "Sheet" not in [_["alias"] for _ in tables] and "Sheet" != "disclaimer_name":
        std=wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)
    # Writing to disk
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    key = uid + "/" + title+".xlsx"
    # key = key.decode("utf-8").encode("ascii","ignore")
    save(key, output)
    return get_signed(key)

def save(key, output):
    obj = s3.Object(Bucket, key)
    obj.put(Body=output)

def get_signed(k):
    client = boto3.client('lambda')
    resp = client.invoke(
        FunctionName=f"s3-{ENV}-signed",
        InvocationType='RequestResponse',
        LogType='Tail',
        Payload=json.dumps({'expires':60, 'key':k, 'bucket':Bucket})
    )
    # The response contains the presigned URL
    return json.loads(resp['Payload'].read())



if __name__== "__main__":
    print("!")
